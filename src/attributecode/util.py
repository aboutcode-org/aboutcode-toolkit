#!/usr/bin/env python
# -*- coding: utf8 -*-
# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
# ============================================================================

from collections import OrderedDict

import codecs
import csv
import json
import ntpath
import openpyxl
import os
import posixpath
import re
import shutil
import string
import sys
from distutils.dir_util import copy_tree
from itertools import zip_longest

from attributecode import CRITICAL
from attributecode import WARNING
from attributecode import Error

on_windows = 'win32' in sys.platform

# boolean field name
boolean_fields = ['redistribute', 'attribute', 'track_change', 'modified', 'internal_use_only']
file_fields = ['about_resource', 'notice_file', 'changelog_file', 'author_file']


def to_posix(path):
    """
    Return a path using the posix path separator given a path that may contain
    posix or windows separators, converting "\\" to "/". NB: this path will
    still be valid in the windows explorer (except for a UNC or share name). It
    will be a valid path everywhere in Python. It will not be valid for windows
    command line operations.
    """
    return path.replace(ntpath.sep, posixpath.sep)


UNC_PREFIX = u'\\\\?\\'
UNC_PREFIX_POSIX = to_posix(UNC_PREFIX)
UNC_PREFIXES = (UNC_PREFIX_POSIX, UNC_PREFIX,)

valid_file_chars = string.digits + string.ascii_letters + '_-.+()~[]{}|@%' + ' '


def invalid_chars(path):
    """
    Return a list of invalid characters in the file name of `path`.
    """
    path = to_posix(path)
    rname = resource_name(path)
    name = rname.lower()
    return [c for c in name if c not in valid_file_chars]


def check_file_names(paths):
    """
    Given a sequence of file paths, check that file names are valid and that
    there are no case-insensitive duplicates in any given directories.
    Return a list of errors.

    From spec :
        A file name can contain only these US-ASCII characters:
        - digits from 0 to 9
        - uppercase and lowercase letters from A to Z
        - the _ underscore, - dash and . period signs.
    From spec:
     The case of a file name is not significant. On case-sensitive file
     systems (such as Linux), a tool must raise an error if two ABOUT files
     stored in the same directory have the same lowercase file name.
    """
    # FIXME: this should be a defaultdicts that accumulates all duplicated paths
    seen = {}
    errors = []
    for orig_path in paths:
        path = orig_path
        invalid = invalid_chars(path)
        if invalid:
            invalid = ''.join(invalid)
            msg = ('Invalid characters %(invalid)r in file name at: '
                   '%(path)r' % locals())
            errors.append(Error(CRITICAL, msg))

        path = to_posix(orig_path)
        name = resource_name(path).lower()
        parent = posixpath.dirname(path)
        path = posixpath.join(parent, name)
        path = posixpath.normpath(path)
        path = posixpath.abspath(path)
        existing = seen.get(path)
        if existing:
            msg = ('Duplicate files: %(orig_path)r and %(existing)r '
                   'have the same case-insensitive file name' % locals())
            errors.append(Error(CRITICAL, msg))
        else:
            seen[path] = orig_path
    return errors


def wrap_boolean_value(context):
    updated_context = ''
    for line in context.splitlines():
        """
        wrap the boolean value in quote
        """
        key = line.partition(':')[0]
        value = line.partition(':')[2].strip()
        value = '"' + value + '"'
        if key in boolean_fields and not value == "":
            updated_context += key + ': ' + value + '\n'
        else:
            updated_context += line + '\n'
    return updated_context


def replace_tab_with_spaces(context):
    updated_context = ''
    for line in context.splitlines():
        """
        Replace tab with 4 spaces
        """
        updated_context += line.replace('\t', '    ') + '\n'
    return updated_context


# TODO: rename to normalize_path
def get_absolute(location):
    """
    Return an absolute normalized location.
    """
    location = os.path.expanduser(location)
    location = os.path.expandvars(location)
    location = os.path.normpath(location)
    location = os.path.abspath(location)
    return location


def get_locations(location):
    """
    Return a list of locations of files given the `location` of a
    a file or a directory tree containing ABOUT files.
    File locations are normalized using posix path separators.
    """
    location = add_unc(location)
    location = get_absolute(location)
    assert os.path.exists(location)

    if os.path.isfile(location):
        yield location
    else:
        for base_dir, _, files in os.walk(location):
            for name in files:
                bd = to_posix(base_dir)
                yield posixpath.join(bd, name)


def get_about_locations(location):
    """
    Return a list of locations of ABOUT files given the `location` of a
    a file or a directory tree containing ABOUT files.
    File locations are normalized using posix path separators.
    """
    for loc in get_locations(location):
        if is_about_file(loc):
            yield loc


def norm(p):
    """
    Normalize the path
    """
    if p.startswith(UNC_PREFIX) or p.startswith(to_posix(UNC_PREFIX)):
        p = p.strip(UNC_PREFIX).strip(to_posix(UNC_PREFIX))
    p = to_posix(p)
    p = p.strip(posixpath.sep)
    p = posixpath.normpath(p)
    return p


def get_relative_path(base_loc, full_loc):
    """
    Return a posix path for a given full location relative to a base location.
    The first segment of the different between full_loc and base_loc will become
    the first segment of the returned path.
    """
    base = norm(base_loc)
    path = norm(full_loc)

    assert path.startswith(base), ('Cannot compute relative path: '
                                   '%(path)r does not start with %(base)r'
                                   % locals())
    base_name = resource_name(base)
    no_dir = base == base_name
    same_loc = base == path
    if same_loc:
        # this is the case of a single file or single dir
        if no_dir:
            # we have no dir: the full path is the same as the resource name
            relative = base_name
        else:
            # we have at least one dir
            parent_dir = posixpath.dirname(base)
            parent_dir = resource_name(parent_dir)
            relative = posixpath.join(parent_dir, base_name)
    else:
        relative = path[len(base) + 1:]
        # We don't want to keep the first segment of the root of the returned path.
        # See https://github.com/nexB/attributecode/issues/276
        # relative = posixpath.join(base_name, relative)
    return relative


def to_native(path):
    """
    Return a path using the current OS path separator given a path that may
    contain posix or windows separators, converting "/" to "\\" on windows
    and "\\" to "/" on posix OSes.
    """
    path = path.replace(ntpath.sep, os.path.sep)
    path = path.replace(posixpath.sep, os.path.sep)
    return path


def is_about_file(path):
    """
    Return True if the path represents a valid ABOUT file name.
    """
    if path:
        path = path.lower()
        return path.endswith('.about') and path != '.about'


def resource_name(path):
    """
    Return the file or directory name from a path.
    """
    path = path.strip()
    path = to_posix(path)
    path = path.rstrip(posixpath.sep)
    _left, right = posixpath.split(path)
    return right.strip()


def load_csv(location):
    """
    Read CSV at `location`, return a list of ordered dictionaries, one
    for each row.
    """
    results = []
    with codecs.open(location, mode='rb', encoding='utf-8-sig',
                     errors='replace') as csvfile:
        for row in csv.DictReader(csvfile):
            # convert all the column keys to lower case
            updated_row = {key.lower().strip(): value for key, value in row.items()}
            results.append(updated_row)
    return results

def load_json(location):
    """
    Read JSON file at `location` and return a list of ordered dicts, one for
    each entry.
    """
    with open(location) as json_file:
        results = json.load(json_file)

    # FIXME: this is too clever and complex... IMHO we should not try to guess the format.
    # instead a command line option should be provided explictly to say what is the format
    if isinstance(results, list):
        results = sorted(results)
    else:
        if u'aboutcode_manager_notice' in results:
            results = results['components']
        elif u'scancode_notice' in results:
            results = results['files']
        else:
            results = [results]
    return results


# FIXME: rename to is_online: BUT do we really need this at all????
def have_network_connection():
    """
    Return True if an HTTP connection to some public web site is possible.
    """
    import socket
    import http.client as httplib

    http_connection = httplib.HTTPConnection('dejacode.org', timeout=10)  # NOQA
    try:
        http_connection.connect()
    except socket.error:
        return False
    else:
        return True


def extract_zip(location):
    """
    Extract a zip file at location in a temp directory and return the temporary
    directory where the archive was extracted.
    """
    import zipfile
    import tempfile

    if not zipfile.is_zipfile(location):
        raise Exception('Incorrect zip file %(location)r' % locals())

    archive_base_name = os.path.basename(location).replace('.zip', '')
    base_dir = tempfile.mkdtemp(prefix='aboutcode-toolkit-extract-')
    target_dir = os.path.join(base_dir, archive_base_name)
    target_dir = add_unc(target_dir)
    os.makedirs(target_dir)

    if target_dir.endswith((ntpath.sep, posixpath.sep)):
        target_dir = target_dir[:-1]

    with zipfile.ZipFile(location) as zipf:
        for info in zipf.infolist():
            name = info.filename
            content = zipf.read(name)
            target = os.path.join(target_dir, name)
            is_dir = target.endswith((ntpath.sep, posixpath.sep))
            if is_dir:
                target = target[:-1]
            parent = os.path.dirname(target)
            if on_windows:
                target = target.replace(posixpath.sep, ntpath.sep)
                parent = parent.replace(posixpath.sep, ntpath.sep)
            if not os.path.exists(parent):
                os.makedirs(add_unc(parent))
            if not content and is_dir:
                if not os.path.exists(target):
                    os.makedirs(add_unc(target))
            if not os.path.exists(target):
                with open(target, 'wb') as f:
                    f.write(content)
    return target_dir


def add_unc(location):
    """
    Convert a `location` to an absolute Window UNC path to support long paths on
    Windows. Return the location unchanged if not on Windows. See
    https://msdn.microsoft.com/en-us/library/aa365247.aspx
    """
    if on_windows and not location.startswith(UNC_PREFIX):
        if location.startswith(UNC_PREFIX_POSIX):
            return UNC_PREFIX + os.path.abspath(location.strip(UNC_PREFIX_POSIX))
        return UNC_PREFIX + os.path.abspath(location)
    return location


# FIXME: add docstring
def copy_license_notice_files(fields, base_dir, reference_dir, afp):
    """
    Given a list of (key, value) `fields` tuples and a `base_dir` where ABOUT
    files and their companion LICENSe are store, and an extra `reference_dir`
    where reference license an notice files are stored and the `afp`
    about_file_path value, this function will copy to the base_dir the
    license_file or notice_file if found in the reference_dir
    """
    errors = []
    copy_file_name = ''
    for key, value in fields:
        if key == 'license_file' or key == 'notice_file':
            if value:
                # This is to handle multiple license_file value in CSV format
                # The following code will construct a list to contain the
                # license file(s) that need to be copied.
                # Note that *ONLY* license_file field allows \n. Others file
                # fields that have \n will prompts error at validation stage
                file_list = []
                if '\n' in value:
                    f_list = value.split('\n')
                else:
                    if not isinstance(value, list):
                        f_list = [value]
                    else:
                        f_list = value
                # The following code is to adopt the approach from #404
                # to use comma for multiple files which refer the same license
                for item in f_list:
                    if ',' in item:
                        item_list = item.split(',')
                        for i in item_list:
                            file_list.append(i.strip())
                    else:
                        file_list.append(item)
            else:
                continue

            for copy_file_name in file_list:
                from_lic_path = posixpath.join(to_posix(reference_dir), copy_file_name)
                about_file_dir = os.path.dirname(to_posix(afp)).lstrip('/')
                to_lic_path = posixpath.join(to_posix(base_dir), about_file_dir)
                if not os.path.exists(posixpath.join(to_lic_path, copy_file_name)):
                    err = copy_file(from_lic_path, to_lic_path)
                    if err:
                        errors.append(err)
    return errors


def copy_file(from_path, to_path):
    error = ''
    # Return if the from_path is empty or None.
    if not from_path:
        return

    if on_windows:
        if not from_path.startswith(UNC_PREFIXES):
            from_path = add_unc(from_path)
        if not to_path.startswith(UNC_PREFIXES):
            to_path = add_unc(to_path)

    # Strip the white spaces
    from_path = from_path.strip()
    to_path = to_path.strip()
    # Errors will be captured when doing the validation
    if not os.path.exists(from_path):
        return ''

    if not posixpath.exists(to_path):
        os.makedirs(to_path)
    try:
        if os.path.isdir(from_path):
            # Copy the whole directory structure
            if from_path.endswith('/'):
                from_path = from_path.rpartition('/')[0]
            folder_name = os.path.basename(from_path)
            to_path = os.path.join(to_path, folder_name)
            if os.path.exists(to_path):
                msg = to_path + ' is already existed and is replaced by ' + from_path
                error = Error(WARNING, msg)
            copy_tree(from_path, to_path)
        else:
            file_name = os.path.basename(from_path)
            to_file_path = os.path.join(to_path, file_name)
            if os.path.exists(to_file_path):
                msg = to_file_path + ' is already existed and is replaced by ' + from_path
                error = Error(WARNING, msg)
            shutil.copy2(from_path, to_path)
        return error
    except Exception as e:
        msg = 'Cannot copy file at %(from_path)r.' % locals()
        error = Error(CRITICAL, msg)
        return error


# FIXME: we should use a license object instead
def ungroup_licenses(licenses):
    """
    Ungroup multiple licenses information
    """
    lic_key = []
    lic_name = []
    lic_file = []
    lic_url = []
    spdx_lic_key = []
    lic_score = []
    for lic in licenses:
        if 'key' in lic:
            lic_key.append(lic['key'])
        if 'name' in lic:
            lic_name.append(lic['name'])
        if 'file' in lic:
            lic_file.append(lic['file'])
        if 'url' in lic:
            lic_url.append(lic['url'])
        if 'spdx_license_key' in lic:
            spdx_lic_key.append(lic['spdx_license_key'])
        if 'score' in lic:
            lic_score.append(lic['score'])
    return lic_key, lic_name, lic_file, lic_url, spdx_lic_key, lic_score


# FIXME: add docstring
def format_about_dict_output(about_dictionary_list):
    formatted_list = []
    for element in about_dictionary_list:
        row_list = dict()
        for key in element:
            if element[key]:
                if isinstance(element[key], list):
                    row_list[key] = u'\n'.join((element[key]))
                elif key == u'about_resource':
                    row_list[key] = u'\n'.join((element[key].keys()))
                else:
                    row_list[key] = element[key]
        formatted_list.append(row_list)
    return formatted_list


# FIXME: add docstring
def format_about_dict_for_json_output(about_dictionary_list):
    licenses = ['license_key', 'license_name', 'license_file', 'license_url']
    json_formatted_list = []
    for element in about_dictionary_list:
        row_list = dict()
        # FIXME: aboid using parallel list... use an object instead
        license_key = []
        license_name = []
        license_file = []
        license_url = []

        for key in element:
            if element[key]:
                # The 'about_resource' is an ordered dict
                if key == 'about_resource':
                    row_list[key] = list(element[key].keys())[0]
                elif key in licenses:
                    if key == 'license_key':
                        license_key = element[key]
                    elif key == 'license_name':
                        license_name = element[key]
                    elif key == 'license_file':
                        license_file = element[key]
                    elif key == 'license_url':
                        license_url = element[key]
                else:
                    row_list[key] = element[key]

        # Group the same license information in a list
        license_group = list(zip_longest(license_key, license_name, license_file, license_url))
        if license_group:
            licenses_list = []
            for lic_group in license_group:
                lic_dict = dict()
                if lic_group[0]:
                    lic_dict['key'] = lic_group[0]
                if lic_group[1]:
                    lic_dict['name'] = lic_group[1]
                if lic_group[2]:
                    lic_dict['file'] = lic_group[2]
                if lic_group[3]:
                    lic_dict['url'] = lic_group[3]
                licenses_list.append(lic_dict)
            row_list['licenses'] = licenses_list
        json_formatted_list.append(row_list)
    return json_formatted_list


def unique(sequence):
    """
    Return a list of unique items found in sequence. Preserve the original
    sequence order.
    For example:
    >>> unique([1, 5, 3, 5])
    [1, 5, 3]
    """
    deduped = []
    for item in sequence:
        if item not in deduped:
            deduped.append(item)
    return deduped


def filter_errors(errors, minimum_severity=WARNING):
    """
    Return a list of unique `errors` Error object filtering errors that have a
    severity below `minimum_severity`.
    """
    return [e for e in errors if e.severity >= minimum_severity]


def create_dir(location):
    """
    Create directory or directory tree at location, ensuring it is readable
    and writeable.
    """
    import stat
    if not os.path.exists(location):
        os.makedirs(location)
        os.chmod(location, stat.S_IRWXU | stat.S_IRWXG
                 | stat.S_IROTH | stat.S_IXOTH)


def get_temp_dir(sub_dir_path=None):
    """
    Create a unique new temporary directory location. Create directories
    identified by sub_dir_path if provided in this temporary directory.
    Return the location for this unique directory joined with the
    sub_dir_path if any.
    """
    new_temp_dir = build_temp_dir()

    if sub_dir_path:
        # create a sub directory hierarchy if requested
        new_temp_dir = os.path.join(new_temp_dir, sub_dir_path)
        create_dir(new_temp_dir)
    return new_temp_dir


def build_temp_dir(prefix='attributecode-'):
    """
    Create and return a new unique empty directory created in base_dir.
    """
    import tempfile
    location = tempfile.mkdtemp(prefix=prefix)
    create_dir(location)
    return location

def get_file_text(file_name, reference):
    """
    Return the file content from the license_file/notice_file field from the
    given reference directory.
    """
    error = ''
    text = ''
    file_path = os.path.join(reference, file_name)
    if not os.path.exists(file_path):
        msg = "The file " + file_path + " does not exist"
        error = Error(CRITICAL, msg)
    else:
        with codecs.open(file_path, 'rb', encoding='utf-8-sig', errors='replace') as txt:
        #with io.open(file_path, encoding='utf-8') as txt:
            text = txt.read()
    return error, text

def convert_object_to_dict(about):
    """
    Convert the list of field object
        [Field(name='name', value=''), Field(name='version', value='')]
    to a dictionary
    """
    about_dict = {}
    # Convert all the supported fields into a dictionary
    fields_dict = getattr(about, 'fields')
    custom_fields_dict = getattr(about, 'custom_fields')
    supported_dict = {**fields_dict, **custom_fields_dict}
    for field in supported_dict:
        key = supported_dict[field].name
        value = supported_dict[field].value
        about_dict[key] = value
    return about_dict

def load_scancode_json(location):
    """
    Read the scancode JSON file at `location` and return a list of dictionaries.
    """
    updated_results = []

    with open(location) as json_file:
        results = json.load(json_file)
    results = results['files']
    # Rename the "path" to "about_resource"
    for item in results:
        updated_dict = {}
        for key in item:
            if key == 'path':
                updated_dict['about_resource'] = item[key]
            else:
                updated_dict[key] = item[key]
        updated_results.append(updated_dict)
    return updated_results

def load_excel(location):
    """
    Read XLSX at `location`, return a list of ordered dictionaries, one
    for each row.
    """
    results = []
    errors = []
    import warnings

    # This is to prevent showing the: warn("Workbook contains no default style, apply openpyxl's default")
    with warnings.catch_warnings(record=True):
        sheet_obj = openpyxl.load_workbook(location).active
    max_col = sheet_obj.max_column

    index = 1
    col_keys = []
    mapping_dict = {}

    while index <= max_col:
        value = sheet_obj.cell(row=1, column=index).value
        if value in col_keys:
            msg = 'Duplicated column name, ' + str(value) + ', detected.' 
            errors.append(Error(CRITICAL, msg))
            return errors, results
        if value in mapping_dict:
            value = mapping_dict[value]
        col_keys.append(value)
        index = index + 1

    for row in sheet_obj.iter_rows(min_row=2, values_only=True):
        row_dict = OrderedDict()
        index = 0
        while index < max_col:
            value = row[index]
            if value:
                row_dict[col_keys[index]] = value
            else:
                row_dict[col_keys[index]] = ''
            index = index + 1
        results.append(row_dict)
    return errors, results

def write_licenses(lic_dict, location):
    import io

    loc = to_posix(location)
    errors = []

    if not posixpath.exists(loc):
        os.makedirs(add_unc(loc))
    try:
        for lic in lic_dict:
            output_location = posixpath.join(loc, lic)
            with io.open(output_location, 'w', encoding='utf-8', errors='replace') as out:
                out.write(lic_dict[lic])
    except Exception as e:
        msg = str(e)
        errors.append(Error(CRITICAL, msg))
    return errors


"""
Return True if a string s  name is safe to use as an attribute name.
"""
is_valid_name = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$').match
