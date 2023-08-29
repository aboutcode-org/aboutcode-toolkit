#!/usr/bin/env python
# -*- coding: utf8 -*-
# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
# ============================================================================

import json
from collections import Counter, OrderedDict
from itertools import zip_longest

import attr
import openpyxl

from attributecode import CRITICAL
from attributecode import Error
from attributecode import saneyaml
from attributecode.util import csv
from attributecode.util import replace_tab_with_spaces


def transform_csv(location):
    """
    Read a CSV file at `location` and convert data into list of dictionaries.
    """
    errors = []
    new_data = []
    rows = read_csv_rows(location)
    data = iter(rows)
    names = next(rows)
    field_names = strip_trailing_fields_csv(names)
    dupes = check_duplicate_fields(field_names)

    if dupes:
        msg = u'Duplicated field name: %(name)s'
        for name in dupes:
            errors.append(Error(CRITICAL, msg % locals()))

    if not errors:
        # Convert to dicts
        new_data = [dict(zip_longest(field_names, item)) for item in data]

    return new_data, errors


def transform_json(location):
    """
    Read a JSON file at `location` and convert data into list of dictionaries.
    """
    errors = []
    new_data = []
    items = read_json(location)
    data = normalize_dict_data(items)
    new_data = strip_trailing_fields_json(data)

    return new_data, errors


def transform_excel(location, worksheet=None):
    """
    Read a XLSX file at `location` and convert data into list of dictionaries.
    """
    errors = []
    new_data = []
    dupes, new_data = read_excel(location, worksheet)
    if dupes:
        msg = u'Duplicated field name: %(name)s'
        for name in dupes:
            errors.append(Error(CRITICAL, msg % locals()))
    return new_data, errors


def strip_trailing_fields_csv(names):
    """
    Strip trailing spaces for field names #456
    """
    field_names = []
    for name in names:
        field_names.append(name.strip())
    return field_names


def strip_trailing_fields_json(items):
    """
    Strip trailing spaces for field name #456
    """
    data = []
    for item in items:
        od = {}
        for field in item:
            stripped_field_name = field.strip()
            od[stripped_field_name] = item[field]
        data.append(od)
    return data


def normalize_dict_data(data):
    """
    Check if the input data from scancode-toolkit and normalize to a normal
    dictionary if it is.
    Return a list type of normalized dictionary.
    """
    try:
        # Check if this is a JSON output from scancode-toolkit
        if (data["headers"][0]["tool_name"] == "scancode-toolkit"):
            # only takes data inside "files"
            new_data = data["files"]
    except:
        new_data = data
    if not isinstance(new_data, list):
        new_data = [new_data]
    return new_data


def transform_data(data, transformer):
    """
    Read a dictionary and apply transformations using the
    `transformer` Transformer.
    Return a tuple of:
       ([field names...], [transformed ordered dict...], [Error objects..])
    """
    renamed_field_data = transformer.apply_renamings(data)

    if transformer.field_filters:
        renamed_field_data = list(
            transformer.filter_fields(renamed_field_data))

    if transformer.exclude_fields:
        renamed_field_data = list(
            transformer.filter_excluded(renamed_field_data))

    errors = transformer.check_required_fields(renamed_field_data)
    if errors:
        return data, errors
    return renamed_field_data, errors


tranformer_config_help = '''
A transform configuration file is used to describe which transformations and
validations to apply to a source CSV file. This is a simple text file using YAML
format, using the same format as an .ABOUT file.

The attributes that can be set in a configuration file are:

* field_renamings:
An optional map of source CSV or JSON field name to target CSV/JSON new field name that
is used to rename CSV fields.

For instance with this configuration the fields "Directory/Location" will be
renamed to "about_resource" and "foo" to "bar":
    field_renamings:
        about_resource : 'Directory/Location'
        bar : foo

The renaming is always applied first before other transforms and checks. All
other field names referenced below are these that exist AFTER the renamings
have been applied to the existing field names.

* required_fields:
An optional list of required field names that must have a value, beyond the
standard fields names. If a source CSV/JSON does not have such a field or a row is
missing a value for a required field, an error is reported.

For instance with this configuration an error will be reported if the fields
"name" and "version" are missing or if any row does not have a value set for
these fields:
    required_fields:
        - name
        - version

* field_filters:
An optional list of field names that should be kept in the transformed CSV/JSON. If
this list is provided, all the fields from the source CSV/JSON that should be kept
in the target CSV/JSON must be listed regardless of  either standard or required
fields. If this list is not provided, all source CSV/JSON fields are kept in the
transformed target CSV/JSON.

For instance with this configuration the target CSV/JSON will only contains the "name"
and "version" fields and no other field:
    field_filters:
        - name
        - version

* exclude_fields:
An optional list of field names that should be excluded in the transformed CSV/JSON. If
this list is provided, all the fields from the source CSV/JSON that should be excluded
in the target CSV/JSON must be listed. Excluding standard or required fields will cause
an error. If this list is not provided, all source CSV/JSON fields are kept in the
transformed target CSV/JSON.

For instance with this configuration the target CSV/JSON will not contain the "type"
and "temp" fields:
    exclude_fields:
        - type
        - temp
'''


@attr.attributes
class Transformer(object):
    __doc__ = tranformer_config_help

    field_renamings = attr.attrib(default=attr.Factory(dict))
    required_fields = attr.attrib(default=attr.Factory(list))
    field_filters = attr.attrib(default=attr.Factory(list))
    exclude_fields = attr.attrib(default=attr.Factory(list))

    # a list of all the standard fields from AboutCode toolkit
    standard_fields = attr.attrib(default=attr.Factory(list), init=False)
    # a list of the subset of standard fields that are essential and MUST be
    # present for AboutCode toolkit to work
    essential_fields = attr.attrib(default=attr.Factory(list), init=False)

    # called by attr after the __init__()
    def __attrs_post_init__(self, *args, **kwargs):
        from attributecode.model import About
        about = About()
        self.essential_fields = list(about.required_fields)
        self.standard_fields = [f.name for f in about.all_fields()]

    @classmethod
    def default(cls):
        """
        Return a default Transformer with built-in transforms.
        """
        return cls(
            field_renamings={},
            required_fields=[],
            field_filters=[],
            exclude_fields=[],
        )

    @classmethod
    def from_file(cls, location):
        """
        Load and return a Transformer instance from a YAML configuration file at
        `location`.
        """
        with open(location, encoding='utf-8', errors='replace') as conf:
            data = saneyaml.load(replace_tab_with_spaces(conf.read()))
        return cls(
            field_renamings=data.get('field_renamings', {}),
            required_fields=data.get('required_fields', []),
            field_filters=data.get('field_filters', []),
            exclude_fields=data.get('exclude_fields', []),
        )

    def check_required_fields(self, data):
        """
        Return a list of Error for a `data` list of ordered dict where a
        dict is missing a value for a required field name.
        """
        errors = []
        required = set(self.essential_fields + self.required_fields)
        if not required:
            return []

        for rn, item in enumerate(data):
            missings = [rk for rk in required if not item.get(rk)]
            if not missings:
                continue

            missings = ', '.join(missings)
            msg = 'Row {rn} is missing required values for fields: {missings}'
            errors.append(Error(CRITICAL, msg.format(**locals())))

        return errors

    def apply_renamings(self, data):
        """
        Return a tranformed list of `field_names` where fields are renamed
        based on this Transformer configuration.
        """
        renamings = self.field_renamings
        renamed_to_list = list(renamings.keys())
        renamed_from_list = list(renamings.values())
        if not renamings:
            return data
        if isinstance(data, dict):
            renamed_obj = {}
            for key, value in data.items():
                if key in renamed_from_list:
                    for idx, renamed_from_key in enumerate(renamed_from_list):
                        if key == renamed_from_key:
                            renamed_key = renamed_to_list[idx]
                            renamed_obj[renamed_key] = self.apply_renamings(
                                value)
                else:
                    renamed_obj[key] = self.apply_renamings(value)
            return renamed_obj
        elif isinstance(data, list):
            return [self.apply_renamings(item) for item in data]
        else:
            return data

    """
    def clean_fields(self, field_names):

        Apply standard cleanups to a list of fields and return these.

        if not field_names:
            return field_names
        return [c.strip().lower() for c in field_names]
    """

    def filter_fields(self, data):
        """
        Yield transformed dicts from a `data` list of dicts keeping only
        fields with a name in the `field_filters`of this Transformer.
        Return the data unchanged if no `field_filters` exists.
        """
        # field_filters = set(self.clean_fields(self.field_filters))
        field_filters = set(self.field_filters)
        for entry in data:
            yield {k: v for k, v in entry.items() if k in field_filters}

    def filter_excluded(self, data):
        """
        Yield transformed dicts from a `data` list of dicts excluding
        fields with names in the `exclude_fields`of this Transformer.
        Return the data unchanged if no `exclude_fields` exists.
        """
        # exclude_fields = set(self.clean_fields(self.exclude_fields))
        exclude_fields = set(self.exclude_fields)
        filtered_list = []
        for entry in data:
            result = {}
            for k, v in entry.items():
                if type(v) == list:
                    result[k] = self.filter_excluded(v)
                elif k not in exclude_fields:
                    result[k] = v
            filtered_list.append(result)
            # yield result
            # yield {k: v for k, v in entry.items() if k not in exclude_fields}
        return filtered_list


def check_duplicate_fields(field_names):
    """
    Check that there are no duplicate in the `field_names` list of field name
    strings, ignoring case. Return a list of unique duplicated field names.
    """
    counted = Counter(c.lower() for c in field_names)
    return [field for field, count in sorted(counted.items()) if count > 1]


def read_csv_rows(location):
    """
    Yield rows (as a list of values) from a CSV file at `location`.
    """
    with open(location, encoding='utf-8', errors='replace') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            yield row


def read_json(location):
    """
    Yield rows (as a list of values) from a CSV file at `location`.
    """
    with open(location, encoding='utf-8', errors='replace') as jsonfile:
        return json.load(jsonfile)


def write_csv(location, data):
    """
    Write a CSV file at `location` with the `data` which is a list of ordered dicts.
    """
    field_names = list(data[0].keys())
    with open(location, 'w', encoding='utf-8', newline='\n', errors='replace') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=field_names)
        writer.writeheader()
        writer.writerows(data)


def write_json(location, data):
    """
    Write a JSON file at `location` the `data` list of ordered dicts.
    """
    with open(location, 'w') as jsonfile:
        json.dump(data, jsonfile, indent=3)


def read_excel(location, worksheet=None):
    """
    Read XLSX at `location`, return a list of ordered dictionaries, one
    for each row.
    """
    results = []
    errors = []
    input_bom = openpyxl.load_workbook(location)
    if worksheet:
        sheet_obj = input_bom[worksheet]
    else:
        sheet_obj = input_bom.active
    max_col = sheet_obj.max_column

    index = 1
    col_keys = []
    mapping_dict = {}
    while index <= max_col:
        value = sheet_obj.cell(row=1, column=index).value
        if value in col_keys:
            msg = 'Duplicated column name, ' + str(value) + ', detected.'
            errors.append(Error(CRITICAL, msg))
            return errors, results
        if value in mapping_dict:
            value = mapping_dict[value]
        col_keys.append(value)
        index = index + 1

    for row in sheet_obj.iter_rows(min_row=2, values_only=True):
        row_dict = OrderedDict()
        index = 0
        while index < max_col:
            value = row[index]
            if value:
                row_dict[col_keys[index]] = value
            else:
                row_dict[col_keys[index]] = ''
            index = index + 1
        results.append(row_dict)
    return errors, results


def write_excel(location, data):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Get the header
    headers = list(data[0].keys())
    ws.append(headers)

    for elements in data:
        ws.append([elements.get(h) for h in headers])

    wb.save(location)
